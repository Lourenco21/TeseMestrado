import csv
import os
import re
import traceback
from datetime import datetime, time
from io import StringIO

import unicodedata
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404
from django.utils.text import slugify
from openpyxl import load_workbook

import pandas as pd
import numpy as np
import requests
from pandas.errors import ParserError
from rest_framework import status, generics
from rest_framework.generics import ListAPIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .problem_schemas import get_problem_schema, PROBLEM_SCHEMAS
from .problem_schemas.problem_catalog import PROBLEM_FAMILIES, CONSTRAINT_LIBRARY, OBJECTIVE_LIBRARY
from .problem_schemas.rooms import ROOMS_FILE_SCHEMA
from .serializers import ScheduleSerializer, ScheduleListSerializer, ProblemDraftSerializer, RoomDataFileSerializer, \
    SolutionListSerializer, SolutionDetailSerializer
from .models import Schedule, ProblemDraft, RoomDataFile, Solution
from .services.file_reader import read_schedule_file, extract_columns_and_preview
from .services.column_matcher import match_canonical_fields_to_source_columns


class ScheduleUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        serializer = ScheduleSerializer(data=request.data)
        if serializer.is_valid():
            schedule = serializer.save()
            return Response(
                {
                    "id": schedule.id,
                    "name": schedule.name,
                    "file": schedule.file.url,
                    "uploaded_at": schedule.uploaded_at,
                    "updated_at": schedule.updated_at,
                },
                status=status.HTTP_201_CREATED,
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProblemMappingSuggestionsView(APIView):
    def get(self, request, problem_id, *args, **kwargs):
        try:
            problem_draft = ProblemDraft.objects.get(pk=problem_id)
        except ProblemDraft.DoesNotExist:
            return Response(
                {"error": "Problem draft not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        if not problem_draft.uploaded_schedule_id:
            return Response(
                {"error": "Problem draft has no uploaded schedule."},
                status=status.HTTP_400_BAD_REQUEST
            )

        schema = get_problem_schema(problem_draft.problem_subtype)
        if not schema:
            return Response(
                {"error": "Schema not found for this problem subtype."},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            schedule = problem_draft.uploaded_schedule
        except Exception:
            return Response(
                {"error": "Schedule not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            df = read_schedule_file(schedule.file.path)
        except Exception as exc:
            return Response(
                {"error": f"Erro ao ler ficheiro: {str(exc)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        file_info = extract_columns_and_preview(df)
        source_columns = file_info["columns"]

        matches = match_canonical_fields_to_source_columns(
            source_columns,
            schema["fields"]
        )

        preview_lookup = {
            item["source_column"]: item["sample_values"]
            for item in file_info["preview"]
        }

        enriched_matches = []
        for match in matches:
            suggested_column = match.get("suggested_source_column") or ""

            enriched_matches.append({
                "target_field": match["target_field"],
                "target_label": match["target_label"],
                "description": next(
                    (field.get("description", "") for field in schema["fields"] if
                     field["key"] == match["target_field"]),
                    ""
                ),
                "required": next(
                    (field.get("required", False) for field in schema["fields"] if
                     field["key"] == match["target_field"]),
                    False
                ),
                "data_type": next(
                    (field.get("data_type", "string") for field in schema["fields"] if
                     field["key"] == match["target_field"]),
                    "string"
                ),
                "aliases": next(
                    (field.get("aliases", []) for field in schema["fields"] if field["key"] == match["target_field"]),
                    []
                ),
                "suggested_source_column": suggested_column,
                "confidence": match.get("confidence", 0.0),
                "match_type": match.get("match_type", "none"),
                "sample_values": preview_lookup.get(suggested_column, []) if suggested_column else [],
                "available_source_columns": source_columns,
            })

        selected_mappings = {
            item["target_field"]: item["suggested_source_column"] or ""
            for item in enriched_matches
        }

        return Response({
            "problem_id": problem_draft.id,
            "problem_family": problem_draft.problem_family,
            "problem_subtype": problem_draft.problem_subtype,
            "schema": schema,
            "source_columns": source_columns,
            "matches": enriched_matches,
            "selected_mappings": selected_mappings,
            "mode": "suggested",
        })


class ScheduleListView(ListAPIView):
    queryset = Schedule.objects.all().order_by("-uploaded_at")
    serializer_class = ScheduleListSerializer


class ProblemDraftListCreateView(generics.ListCreateAPIView):
    queryset = ProblemDraft.objects.all().order_by("-updated_at")
    serializer_class = ProblemDraftSerializer


class ProblemDraftDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProblemDraft.objects.all()
    serializer_class = ProblemDraftSerializer


class ProblemCatalogView(APIView):
    def get(self, request, *args, **kwargs):
        return Response({
            "problem_families": PROBLEM_FAMILIES,
            "objective_library": OBJECTIVE_LIBRARY,
            "constraint_library": CONSTRAINT_LIBRARY,
        })


class RoomDataFileUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        serializer = RoomDataFileSerializer(data=request.data)

        if serializer.is_valid():
            room_file = serializer.save()
            return Response(
                {
                    "id": room_file.id,
                    "name": room_file.name,
                    "file": room_file.file.url,
                    "uploaded_at": room_file.uploaded_at,
                    "updated_at": room_file.updated_at,
                },
                status=status.HTTP_201_CREATED,
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProblemRoomsFilePreviewView(APIView):
    def get(self, request, problem_id, *args, **kwargs):
        try:
            problem_draft = ProblemDraft.objects.get(pk=problem_id)
        except ProblemDraft.DoesNotExist:
            return Response(
                {"error": "Problem draft not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        if not problem_draft.uploaded_rooms_file_id:
            return Response(
                {"error": "Problem draft has no uploaded rooms.py file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            rooms_file = problem_draft.uploaded_rooms_file
        except Exception:
            return Response(
                {"error": "Rooms file not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            df = read_schedule_file(rooms_file.file.path)
        except Exception as exc:
            return Response(
                {"error": f"Erro ao ler ficheiro de salas: {str(exc)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        file_info = extract_columns_and_preview(df)

        schedule_columns = []
        if problem_draft.uploaded_schedule_id:
            try:
                schedule_df = read_schedule_file(problem_draft.uploaded_schedule.file.path)
                schedule_info = extract_columns_and_preview(schedule_df)
                schedule_columns = schedule_info["columns"]
            except Exception:
                schedule_columns = []

        return Response({
            "problem_id": problem_draft.id,
            "rooms_file_id": rooms_file.id,
            "rooms_file_name": rooms_file.name,
            "rooms_source_columns": file_info["columns"],
            "rooms_preview": file_info["preview"],
            "schedule_source_columns": schedule_columns,
            "saved_mapping": problem_draft.rooms_mapping_data or {},
        })


class ProblemRoomsMappingSuggestionsView(APIView):
    def get(self, request, problem_id, *args, **kwargs):
        try:
            problem_draft = ProblemDraft.objects.get(pk=problem_id)
        except ProblemDraft.DoesNotExist:
            return Response(
                {"error": "Problem draft not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        if not problem_draft.uploaded_rooms_file_id:
            return Response(
                {"error": "Problem draft has no uploaded rooms file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            rooms_file = problem_draft.uploaded_rooms_file
        except Exception:
            return Response(
                {"error": "Rooms file not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            df = read_schedule_file(rooms_file.file.path)
        except Exception as exc:
            return Response(
                {"error": f"Erro ao ler ficheiro de salas: {str(exc)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        file_info = extract_columns_and_preview(df)
        source_columns = file_info["columns"]

        schema = ROOMS_FILE_SCHEMA
        matches = match_canonical_fields_to_source_columns(
            source_columns,
            schema["fields"]
        )

        preview_lookup = {
            item["source_column"]: item["sample_values"]
            for item in file_info["preview"]
        }

        enriched_matches = []
        for match in matches:
            suggested_column = match.get("suggested_source_column") or ""
            enriched_matches.append({
                "target_field": match["target_field"],
                "target_label": match["target_label"],
                "description": next(
                    (field.get("description", "") for field in schema["fields"] if
                     field["key"] == match["target_field"]),
                    ""
                ),
                "required": next(
                    (field.get("required", False) for field in schema["fields"] if
                     field["key"] == match["target_field"]),
                    False
                ),
                "data_type": next(
                    (field.get("data_type", "string") for field in schema["fields"] if
                     field["key"] == match["target_field"]),
                    "string"
                ),
                "aliases": next(
                    (field.get("aliases", []) for field in schema["fields"] if field["key"] == match["target_field"]),
                    []
                ),
                "suggested_source_column": suggested_column,
                "confidence": match.get("confidence", 0.0),
                "match_type": match.get("match_type", "none"),
                "sample_values": preview_lookup.get(suggested_column, []) if suggested_column else [],
                "available_source_columns": source_columns,
            })

        selected_field_mappings = {
            item["target_field"]: item["suggested_source_column"] or ""
            for item in enriched_matches
        }

        schedule_columns = []
        suggested_schedule_room_column = ""
        if problem_draft.uploaded_schedule_id:
            try:
                schedule_df = read_schedule_file(problem_draft.uploaded_schedule.file.path)
                schedule_info = extract_columns_and_preview(schedule_df)
                schedule_columns = schedule_info["columns"]
                saved_mapping = (problem_draft.mapping_data or {}).get("mapping", {})
                suggested_schedule_room_column = saved_mapping.get("sala", "")
            except Exception:
                schedule_columns = []

        return Response({
            "problem_id": problem_draft.id,
            "schema": schema,
            "rooms_source_columns": source_columns,
            "rooms_preview": file_info["preview"],
            "matches": enriched_matches,
            "selected_field_mappings": selected_field_mappings,
            "schedule_source_columns": schedule_columns,
            "suggested_schedule_room_column": suggested_schedule_room_column,
            "saved_rooms_mapping_data": problem_draft.rooms_mapping_data or {},
        })


class ProblemRoomsMappingSaveView(APIView):
    def post(self, request, problem_id, *args, **kwargs):
        try:
            problem_draft = ProblemDraft.objects.get(pk=problem_id)
        except ProblemDraft.DoesNotExist:
            return Response(
                {"error": "Problem draft not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        payload = request.data if isinstance(request.data, dict) else {}

        field_mappings = payload.get("field_mappings", {})
        linking = payload.get("linking", {})
        characteristics = payload.get("characteristics", {})

        if not isinstance(field_mappings, dict):
            return Response(
                {"error": "O campo 'field_mappings' tem de ser um objeto JSON."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not isinstance(linking, dict):
            return Response(
                {"error": "O campo 'linking' tem de ser um objeto JSON."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not isinstance(characteristics, dict):
            return Response(
                {"error": "O campo 'characteristics' tem de ser um objeto JSON."},
                status=status.HTTP_400_BAD_REQUEST
            )

        required_field_keys = ["room_name"]
        missing_required = [
            key for key in required_field_keys
            if not field_mappings.get(key)
        ]
        if missing_required:
            return Response(
                {"error": f"Campos obrigatórios em falta: {', '.join(missing_required)}."},
                status=status.HTTP_400_BAD_REQUEST
            )

        schedule_room_column = linking.get("schedule_room_column", "")
        rooms_file_room_column = linking.get("rooms_file_room_column", "")

        if not schedule_room_column or not rooms_file_room_column:
            return Response(
                {"error": "As colunas de ligação são obrigatórias."},
                status=status.HTTP_400_BAD_REQUEST
            )

        char_format = characteristics.get("format", "")
        char_config = characteristics.get("config", {})

        valid_formats = [
            "single_column_list",
            "multiple_columns",
            "range_columns",
        ]
        if char_format not in valid_formats:
            return Response(
                {"error": "Formato de características inválido."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not isinstance(char_config, dict):
            return Response(
                {"error": "O campo 'characteristics.config' tem de ser um objeto JSON."},
                status=status.HTTP_400_BAD_REQUEST
            )

        normalized_config = dict(char_config)

        if char_format == "single_column_list":
            source_column = str(char_config.get("source_column", "")).strip()
            separator = str(char_config.get("separator", "")).strip()

            if not source_column:
                return Response(
                    {"error": "Seleciona a coluna de características."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if not separator:
                return Response(
                    {"error": "Indica o separador das características."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            normalized_config["source_column"] = source_column
            normalized_config["separator"] = separator

        elif char_format == "multiple_flag_columns":
            selected_columns = char_config.get("selected_columns", [])
            selected_values = char_config.get("selected_values")
            active_values = char_config.get("active_values")

            effective_values = selected_values if selected_values is not None else active_values

            if not isinstance(selected_columns, list) or len(selected_columns) == 0:
                return Response(
                    {"error": "Seleciona pelo menos uma coluna de características."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if not isinstance(effective_values, list) or len(effective_values) == 0:
                return Response(
                    {"error": "Indica os valores que significam ativo."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            selected_columns = [
                str(column).strip()
                for column in selected_columns
                if str(column).strip()
            ]
            effective_values = [
                str(value).strip()
                for value in effective_values
                if str(value).strip()
            ]

            if len(selected_columns) == 0:
                return Response(
                    {"error": "Seleciona pelo menos uma coluna de características."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if len(effective_values) == 0:
                return Response(
                    {"error": "Indica os valores que significam ativo."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            normalized_config["selected_columns"] = list(dict.fromkeys(selected_columns))
            normalized_config["selected_values"] = list(dict.fromkeys(effective_values))

        elif char_format == "range_flag_columns":
            start_column = str(char_config.get("start_column", "")).strip()
            end_column = str(char_config.get("end_column", "")).strip()
            selected_values = char_config.get("selected_values")
            active_values = char_config.get("active_values")

            effective_values = selected_values if selected_values is not None else active_values

            if not start_column or not end_column:
                return Response(
                    {"error": "Indica a coluna inicial e final do intervalo."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if not isinstance(effective_values, list) or len(effective_values) == 0:
                return Response(
                    {"error": "Indica os valores que significam ativo."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            effective_values = [
                str(value).strip()
                for value in effective_values
                if str(value).strip()
            ]

            if len(effective_values) == 0:
                return Response(
                    {"error": "Indica os valores que significam ativo."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            normalized_config["start_column"] = start_column
            normalized_config["end_column"] = end_column
            normalized_config["selected_values"] = list(dict.fromkeys(effective_values))

        normalized_characteristics = {
            "format": char_format,
            "config": normalized_config,
        }

        problem_draft.rooms_mapping_data = {
            "field_mappings": field_mappings,
            "linking": linking,
            "characteristics": normalized_characteristics,
        }
        problem_draft.save(update_fields=["rooms_mapping_data", "updated_at"])

        return Response({
            "message": "Mapping das salas guardado com sucesso.",
            "rooms_mapping_data": problem_draft.rooms_mapping_data,
        })


class ProblemSendToJavaView(APIView):
    JAVA_BACKEND_URL = "http://localhost:8080/api/problems/run"

    VALID_RESOLUTION_SCOPES = {
        "semester",
        "week",
        "day",
        "start_half_hour",
    }

    VALID_REPEATED_INSTANCE_STRATEGIES = {
        "reuse_solution",
        "generate_new",
    }

    def post(self, request, problem_id, *args, **kwargs):
        try:
            problem_draft = ProblemDraft.objects.get(pk=problem_id)
        except ProblemDraft.DoesNotExist:
            return Response(
                {"error": "Problem draft not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        resolution_scope = request.data.get("resolution_scope")
        repeated_instance_strategy = request.data.get("repeated_instance_strategy")

        if resolution_scope not in self.VALID_RESOLUTION_SCOPES:
            return Response(
                {"error": "resolution_scope inválido."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if resolution_scope == "semester":
            repeated_instance_strategy = None
        else:
            if repeated_instance_strategy not in self.VALID_REPEATED_INSTANCE_STRATEGIES:
                return Response(
                    {"error": "repeated_instance_strategy inválido para o nível selecionado."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        schedule_row_count = get_file_row_count(
            problem_draft.uploaded_schedule.file if problem_draft.uploaded_schedule else None
        )
        rooms_row_count = get_file_row_count(
            problem_draft.uploaded_rooms_file.file if problem_draft.uploaded_rooms_file else None
        )

        payload = {
            "problem_id": problem_draft.id,
            "name": problem_draft.name,
            "problem_type": problem_draft.problem_family,
            "problem_subtype": problem_draft.problem_subtype,
            "schedule_file_id": problem_draft.uploaded_schedule_id,
            "schedule_file_row_count": schedule_row_count,
            "rooms_file_id": problem_draft.uploaded_rooms_file_id or {},
            "rooms_file_row_count": rooms_row_count or {},
            "mapping_data": problem_draft.mapping_data or {},
            "rooms_mapping_data": problem_draft.rooms_mapping_data or {},
            "constraints": problem_draft.selected_constraints or [],
            "resolution_scope": resolution_scope,
            "repeated_instance_strategy": repeated_instance_strategy,
        }

        missing_fields = []

        if not payload["name"]:
            missing_fields.append("name")
        if not payload["problem_type"]:
            missing_fields.append("problem_type")
        if not payload["problem_subtype"]:
            missing_fields.append("problem_subtype")
        if not payload["schedule_file_id"]:
            missing_fields.append("schedule_file_id")

        if missing_fields:
            return Response(
                {
                    "error": "O problema ainda não está pronto para ser enviado para o Java.",
                    "missing_fields": missing_fields,
                    "payload_preview": payload,
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            java_response = requests.post(
                self.JAVA_BACKEND_URL,
                json=payload,
                timeout=(10,6000),
            )
        except requests.exceptions.RequestException as exc:
            return Response(
                {
                    "error": "Não foi possível contactar o backend Java.",
                    "details": str(exc),
                    "payload_preview": payload,
                },
                status=status.HTTP_502_BAD_GATEWAY
            )

        try:
            java_data = java_response.json()
        except ValueError:
            java_data = {"raw_response": java_response.text}

        if not java_response.ok:
            return Response(
                {
                    "error": "O backend Java respondeu com erro.",
                    "java_status_code": java_response.status_code,
                    "java_response": java_data,
                    "payload_preview": payload,
                },
                status=status.HTTP_502_BAD_GATEWAY
            )

        return Response(
            {
                "message": "Problema enviado com sucesso para o backend Java.",
                "payload_sent": payload,
                "java_response": java_data,
            },
            status=status.HTTP_200_OK
        )


def get_file_row_count(file_field):
    if not file_field:
        return None

    file_path = file_field.path
    extension = os.path.splitext(file_path)[1].lower()

    if extension == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            row_count = sum(1 for row in reader if any(cell.strip() for cell in row)) - 1
        return row_count

    if extension in [".xlsx", ".xlsm"]:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = workbook.active
        row_count = sheet.max_row - 1
        workbook.close()
        return row_count

    return None


class ProblemRequestAlgorithmsView(APIView):
    JAVA_BACKEND_URL = "http://localhost:8080/api/problems/request-algorithms"

    VALID_RESOLUTION_SCOPES = {
        "semester",
        "week",
        "day",
        "start_half_hour",
    }

    VALID_REPEATED_INSTANCE_STRATEGIES = {
        "reuse_solution",
        "generate_new",
    }

    def post(self, request, problem_id, *args, **kwargs):
        try:
            problem_draft = ProblemDraft.objects.get(pk=problem_id)
        except ProblemDraft.DoesNotExist:
            return Response(
                {"error": "Problem draft not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        resolution_scope = request.data.get("resolution_scope")
        repeated_instance_strategy = request.data.get("repeated_instance_strategy")

        if resolution_scope not in self.VALID_RESOLUTION_SCOPES:
            return Response(
                {"error": "resolution_scope inválido."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if resolution_scope == "semester":
            repeated_instance_strategy = None
        elif repeated_instance_strategy not in self.VALID_REPEATED_INSTANCE_STRATEGIES:
            return Response(
                {"error": "repeated_instance_strategy inválido para o nível selecionado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not problem_draft.uploaded_schedule:
            return Response(
                {"error": "O problema não tem ficheiro de horário associado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        missing_fields = []

        if not problem_draft.name:
            missing_fields.append("name")
        if not problem_draft.problem_family:
            missing_fields.append("problem_family")
        if not problem_draft.problem_subtype:
            missing_fields.append("problem_subtype")
        if not problem_draft.mapping_data:
            missing_fields.append("mapping_data")
        if not problem_draft.uploaded_schedule_id:
            missing_fields.append("uploaded_schedule")

        if missing_fields:
            return Response(
                {
                    "error": "O problema ainda não está pronto para pedir algoritmos.",
                    "missing_fields": missing_fields,
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            df = load_schedule_dataframe(problem_draft.uploaded_schedule.file.path)
            mapping = (problem_draft.mapping_data or {}).get("mapping", {}) or {}

            analysis = analyze_schedule_dataframe(df, mapping, resolution_scope)
            constraints_summary = build_constraints_summary(
                problem_draft.selected_constraints or []
            )

            payload = {
                "problem_id": problem_draft.id,
                "name": problem_draft.name,
                "problem_type": problem_draft.problem_family,
                "problem_subtype": problem_draft.problem_subtype,
                "resolution_scope": resolution_scope,
                "repeated_instance_strategy": repeated_instance_strategy,
                "constraints_summary": constraints_summary,
                "instance_characteristics": {
                    "total_classes": int(len(df)),
                    "selected_partition_statistics": analysis["selected_partition_statistics"],
                },
            }

        except Exception as exc:
            return Response(
                {
                    "error": "Não foi possível analisar o ficheiro de horários.",
                    "details": str(exc),
                    "trace": traceback.format_exc(),
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            java_response = requests.post(
                self.JAVA_BACKEND_URL,
                json=payload,
                timeout=(10,6000),
            )
        except requests.exceptions.RequestException as exc:
            return Response(
                {
                    "error": "Não foi possível contactar o backend Java.",
                    "details": str(exc),
                    "payload_preview": payload,
                },
                status=status.HTTP_502_BAD_GATEWAY
            )

        try:
            java_data = java_response.json()
        except ValueError:
            java_data = {"raw_response": java_response.text}

        if not java_response.ok:
            return Response(
                {
                    "error": "O backend Java respondeu com erro.",
                    "java_status_code": java_response.status_code,
                    "java_response": java_data,
                    "payload_preview": payload,
                },
                status=status.HTTP_502_BAD_GATEWAY
            )

        return Response(
            {
                "message": "Pedido de algoritmos enviado com sucesso para o backend Java.",
                "payload_sent": payload,
                "java_response": java_data,
            },
            status=status.HTTP_200_OK
        )



def build_schedule_csv_content(schedule_rows):
    if not isinstance(schedule_rows, list) or not schedule_rows:
        return None

    normalized_rows = [
        row if isinstance(row, dict) else {"value": row}
        for row in schedule_rows
    ]

    fieldnames = []
    for row in normalized_rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)

    buffer = StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=fieldnames,
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()

    for row in normalized_rows:
        writer.writerow({key: row.get(key, "") for key in fieldnames})

    return ContentFile(buffer.getvalue().encode("utf-8"))


class ProblemExecuteView(APIView):
    JAVA_BACKEND_URL = "http://localhost:8080/api/problems/execute"

    VALID_RESOLUTION_SCOPES = {
        "semester",
        "week",
        "day",
        "start_half_hour",
    }

    VALID_REPEATED_INSTANCE_STRATEGIES = {
        "reuse_solution",
        "generate_new",
    }

    def post(self, request, problem_id, *args, **kwargs):
        try:
            problem_draft = ProblemDraft.objects.get(pk=problem_id)
        except ProblemDraft.DoesNotExist:
            return Response(
                {"error": "Problem draft not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        resolution_scope = request.data.get("resolution_scope")
        repeated_instance_strategy = request.data.get("repeated_instance_strategy")
        selected_algorithm_name = request.data.get("selected_algorithm")

        if resolution_scope not in self.VALID_RESOLUTION_SCOPES:
            return Response(
                {"error": "resolution_scope inválido."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if resolution_scope == "semester":
            repeated_instance_strategy = None
        elif repeated_instance_strategy not in self.VALID_REPEATED_INSTANCE_STRATEGIES:
            return Response(
                {"error": "repeated_instance_strategy inválido para o nível selecionado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not selected_algorithm_name:
            return Response(
                {"error": "É obrigatório indicar o algoritmo selecionado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not problem_draft.uploaded_schedule:
            return Response(
                {"error": "O problema não tem ficheiro de horário associado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        missing_fields = []

        if not problem_draft.name:
            missing_fields.append("name")
        if not problem_draft.problem_family:
            missing_fields.append("problem_family")
        if not problem_draft.problem_subtype:
            missing_fields.append("problem_subtype")
        if not problem_draft.mapping_data:
            missing_fields.append("mapping_data")
        if not problem_draft.uploaded_schedule_id:
            missing_fields.append("uploaded_schedule")
        if not problem_draft.selected_constraints:
            missing_fields.append("selected_constraints")

        if missing_fields:
            return Response(
                {
                    "error": "O problema ainda não está pronto para execução.",
                    "missing_fields": missing_fields,
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            df = load_schedule_dataframe(problem_draft.uploaded_schedule.file.path)

            df = apply_room_feature_resolution(
                df,
                problem_draft.mapping_data,
                problem_draft.room_feature_resolution or {},
            )

            mapping = (problem_draft.mapping_data or {}).get("mapping", {}) or {}

            analysis = analyze_schedule_dataframe(df, mapping, resolution_scope)
            constraints_summary = build_constraints_summary(
                problem_draft.selected_constraints or []
            )

            payload = {
                "problem_id": problem_draft.id,
                "name": problem_draft.name,
                "resolution_scope": resolution_scope,
                "repeated_instance_strategy": repeated_instance_strategy,
                "selected_algorithm": selected_algorithm_name,
                "constraints_summary": constraints_summary,
                "selected_constraints": problem_draft.selected_constraints or [],
                "room_feature_resolution": problem_draft.room_feature_resolution or {},
                "resolved_requested_room_features": (
                    df["_resolved_requested_room_features"].tolist()
                    if "_resolved_requested_room_features" in df.columns
                    else []
                ),
                "instance_characteristics": {
                    "total_classes": int(len(df)),
                    "selected_partition_statistics": analysis["selected_partition_statistics"],
                },
                "problem_type": problem_draft.problem_family,
                "problem_subtype": problem_draft.problem_subtype,
                "schedule_id": problem_draft.uploaded_schedule_id,
                "rooms_id": getattr(problem_draft, "rooms_file_id", None),
                "mapping_data": problem_draft.mapping_data,
                "rooms_mapping_data": getattr(problem_draft, "rooms_mapping_data", None),
            }
            print(payload)

        except Exception as exc:
            return Response(
                {
                    "error": "Não foi possível preparar o payload de execução.",
                    "details": str(exc),
                    "trace": traceback.format_exc(),
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            java_response = requests.post(
                self.JAVA_BACKEND_URL,
                json=payload,
                timeout=(10, 6000),
            )
        except requests.exceptions.RequestException as exc:
            return Response(
                {
                    "error": "Não foi possível contactar o backend Java.",
                    "details": str(exc),
                    "payload_preview": payload,
                },
                status=status.HTTP_502_BAD_GATEWAY
            )

        try:
            java_data = java_response.json()
        except ValueError:
            java_data = {"raw_response": java_response.text}

        if not java_response.ok:
            return Response(
                {
                    "error": "O backend Java respondeu com erro na execução.",
                    "java_status_code": java_response.status_code,
                    "java_response": java_data,
                    "payload_preview": payload,
                },
                status=status.HTTP_502_BAD_GATEWAY
            )

        try:
            solution = Solution.objects.create(
                problem=problem_draft,
                status="completed",
                algorithm_used=java_data.get("algorithm_used", selected_algorithm_name),
                used_parameters=java_data.get("used_parameters", {}),
                partition_type=java_data.get("partition_type", ""),
                reuse_solution=bool(java_data.get("reuse_solution", False)),
                constraint_values=java_data.get("constraint_values", {}),
                penalty_summary=java_data.get("penalty_summary", {}),
                partition_count=int(java_data.get("partition_count", 0) or 0),
                execution_time_seconds=float(java_data.get("execution_time_seconds", 0.0)),
            )

            schedule_rows = java_data.get("schedule", [])
            csv_content = build_schedule_csv_content(schedule_rows)

            if csv_content:
                filename = (
                    f"problem-{problem_draft.id}-solution-{solution.id}-"
                    f"{slugify(solution.algorithm_used or 'algorithm')}.csv"
                )
                solution.schedule_file.save(filename, csv_content, save=True)

            serialized_solution = SolutionDetailSerializer(
                solution,
                context={"request": request}
            ).data

        except Exception as exc:
            return Response(
                {
                    "error": "A execução foi concluída, mas não foi possível guardar a solução.",
                    "details": str(exc),
                    "java_response": java_data,
                    "trace": traceback.format_exc(),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(
            {
                "message": "Execução enviada com sucesso para o backend Java.",
                "payload_sent": payload,
                "java_response": java_data,
                "solution": serialized_solution,
            },
            status=status.HTTP_200_OK
        )


def load_schedule_dataframe(file_path):
    lower_path = file_path.lower()

    if lower_path.endswith(".csv"):
        attempts = [
            {"encoding": "utf-8", "sep": ","},
            {"encoding": "utf-8-sig", "sep": ","},
            {"encoding": "latin1", "sep": ","},
            {"encoding": "utf-8", "sep": ";"},
            {"encoding": "utf-8-sig", "sep": ";"},
            {"encoding": "latin1", "sep": ";"},
        ]

        last_error = None

        for attempt in attempts:
            try:
                df = pd.read_csv(
                    file_path,
                    encoding=attempt["encoding"],
                    sep=attempt["sep"],
                    engine="python",
                )
                if df.shape[1] > 1:
                    return df
            except Exception as exc:
                last_error = exc

        raise ValueError(f"Não foi possível ler o CSV. Último erro: {last_error}")

    if lower_path.endswith(".xlsx"):
        try:
            return pd.read_excel(file_path, engine="openpyxl")
        except Exception as exc:
            raise ValueError(f"Erro ao ler XLSX: {exc}")

    if lower_path.endswith(".xls"):
        try:
            return pd.read_excel(file_path)
        except Exception as exc:
            raise ValueError(f"Erro ao ler XLS: {exc}")

    raise ValueError("Formato de ficheiro não suportado.")


def build_constraints_summary(selected_constraints):
    enabled = [item for item in selected_constraints if item.get("enabled", True)]
    hard = [item for item in enabled if item.get("goal") == "hard"]
    soft = [item for item in enabled if item.get("goal") == "soft"]

    return {
        "total": len(enabled),
        "hard": len(hard),
        "soft": len(soft),
        "selected": [
            {
                "id": item.get("id"),
                "goal": item.get("goal"),
                "importance": item.get("importance"),
            }
            for item in enabled
        ],
    }


def analyze_schedule_dataframe(df, mapping, resolution_scope):
    week_column = get_mapped_column(df, mapping, "semana")
    day_column = get_mapped_column(df, mapping, "dia")
    weekday_column = get_mapped_column(df, mapping, "dia_da_semana")
    start_time_column = get_mapped_column(df, mapping, "hora_inicio")

    working_df = df.copy()

    working_df["_parsed_day"] = (
        working_df[day_column].apply(parse_date_value)
        if day_column
        else pd.Series([None] * len(working_df), index=working_df.index)
    )

    working_df["_parsed_start_time"] = (
        working_df[start_time_column].apply(parse_time_value)
        if start_time_column
        else pd.Series([None] * len(working_df), index=working_df.index)
    )

    if resolution_scope == "semester":
        partition_series = pd.Series(["semester"] * len(working_df), index=working_df.index)
    elif resolution_scope == "week":
        partition_series = build_week_partition_series(working_df, week_column)
    elif resolution_scope == "day":
        partition_series = build_day_partition_series(
            working_df, day_column, weekday_column
        )
    elif resolution_scope == "start_half_hour":
        partition_series = build_day_time_partition_series(working_df)
    else:
        partition_series = pd.Series(["unknown"] * len(working_df), index=working_df.index)

    return {
        "selected_partition_statistics": summarize_partitions(partition_series),
    }


def get_mapped_column(df, mapping, logical_key):
    column_name = (mapping or {}).get(logical_key)

    if not column_name:
        return None

    if column_name not in df.columns:
        return None

    return column_name


def build_week_partition_series(df, week_column):
    if week_column and week_column in df.columns:
        week_values = df[week_column].dropna().astype(str).str.strip()

        if not week_values.empty:
            unique_values = set(v.lower() for v in week_values.unique())

            weekday_like_values = {
                "seg", "segunda", "segunda-feira",
                "ter", "terça", "terca", "terça-feira", "terca-feira",
                "qua", "quarta", "quarta-feira",
                "qui", "quinta", "quinta-feira",
                "sex", "sexta", "sexta-feira",
                "sáb", "sab", "sábado", "sabado",
                "dom", "domingo",
                "mon", "monday", "tue", "tuesday", "wed", "wednesday",
                "thu", "thursday", "fri", "friday", "sat", "saturday",
                "sun", "sunday",
            }

            looks_like_weekday_column = unique_values.issubset(weekday_like_values)

            if not looks_like_weekday_column:
                return df[week_column].fillna("unknown_week").astype(str)

    if "_parsed_day" not in df.columns:
        return pd.Series([None] * len(df), index=df.index, dtype="object")

    parsed_days = pd.to_datetime(df["_parsed_day"], errors="coerce").dt.normalize()
    valid_mask = parsed_days.notna()

    result = pd.Series([None] * len(df), index=df.index, dtype="object")

    if not valid_mask.any():
        return result

    min_day = parsed_days.loc[valid_mask].min()
    week_numbers = ((parsed_days.loc[valid_mask] - min_day).dt.days // 7) + 1

    result.loc[valid_mask] = week_numbers.apply(lambda x: f"week_{int(x)}")

    return result


def build_day_partition_series(df, day_column, weekday_column):
    if day_column and "_parsed_day" in df.columns:
        parsed_days = pd.to_datetime(df["_parsed_day"], errors="coerce")
        if parsed_days.notna().any():
            return parsed_days.dt.strftime("%Y-%m-%d").fillna("unknown_day")

    if weekday_column:
        return df[weekday_column].fillna("unknown_day").astype(str)

    return pd.Series(["unknown_day"] * len(df), index=df.index)


def summarize_partitions(partition_series):
    valid_partitions = partition_series.dropna()

    if valid_partitions.empty:
        return {
            "partition_count": 0,
            "average_classes_per_partition": 0,
            "min_classes_per_partition": 0,
            "max_classes_per_partition": 0,
        }

    counts = valid_partitions.astype(str).value_counts()

    return {
        "partition_count": int(counts.shape[0]),
        "average_classes_per_partition": round(float(counts.mean()), 2),
        "min_classes_per_partition": int(counts.min()),
        "max_classes_per_partition": int(counts.max()),
    }


def parse_date_value(value):
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().date()

    if isinstance(value, datetime):
        return value.date()

    if hasattr(value, "date") and not isinstance(value, str):
        try:
            return value.date()
        except Exception:
            pass

    text = str(value).strip()
    if not text:
        return None

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)

    if pd.isna(parsed):
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=False)

    if pd.isna(parsed):
        return None

    return parsed.date()


def parse_time_value(value):
    return coerce_to_time(value)


def time_to_half_hour_label(value):
    parsed_time = coerce_to_time(value)

    if not parsed_time:
        return "unknown_start_time"

    minute = parsed_time.minute
    hour = parsed_time.hour

    if minute >= 30:
        return f"{hour:02d}:30"

    return f"{hour:02d}:00"


def coerce_to_time(value):
    if value is None or pd.isna(value):
        return None

    if isinstance(value, time):
        return value

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().time()

    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, (int, float)):
        try:
            total_seconds = int(round(float(value) * 24 * 60 * 60))
            hours = (total_seconds // 3600) % 24
            minutes = (total_seconds % 3600) // 60
            return time(hour=hours, minute=minutes)
        except Exception:
            return None

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%H:%M", "%H:%M:%S", "%H.%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue

    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None

    return parsed.to_pydatetime().time()


def build_day_time_partition_series(df):
    if "_parsed_day" not in df.columns or "_parsed_start_time" not in df.columns:
        return pd.Series([None] * len(df), index=df.index, dtype="object")

    parsed_days = pd.to_datetime(df["_parsed_day"], errors="coerce").dt.normalize()

    def normalize_half_hour(value):
        parsed_time = coerce_to_time(value)
        if not parsed_time:
            return None

        minute = 30 if parsed_time.minute >= 30 else 0
        return f"{parsed_time.hour:02d}:{minute:02d}"

    normalized_slots = df["_parsed_start_time"].apply(normalize_half_hour)

    result = pd.Series([None] * len(df), index=df.index, dtype="object")

    valid_mask = parsed_days.notna() & normalized_slots.notna()

    result.loc[valid_mask] = (
        parsed_days.loc[valid_mask].dt.strftime("%Y-%m-%d")
        + " "
        + normalized_slots.loc[valid_mask]
    )

    return result


class FilesForJavaView(APIView):
    def get(self, request, problem_id, *args, **kwargs):
        draft = get_object_or_404(ProblemDraft, pk=problem_id)

        if not draft.uploaded_schedule:
            return Response(
                {"detail": "O ProblemDraft não tem horário associado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not draft.uploaded_rooms_file:
            return Response(
                {"detail": "O ProblemDraft não tem ficheiro de salas associado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            schedule_df = read_tabular_file(draft.uploaded_schedule.file.path)
            rooms_df = read_tabular_file(draft.uploaded_rooms_file.file.path)
        except FileNotFoundError as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_404_NOT_FOUND,
            )
        except ValueError as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as exc:
            return Response(
                {
                    "detail": "Erro ao processar os ficheiros do problema.",
                    "error": str(exc),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        schedule_data = build_schedule_data(schedule_df)
        rooms_data = build_rooms_data(rooms_df)

        return Response(
            {
                "schedule_data": schedule_data,
                "rooms_data": rooms_data,
            },
            status=status.HTTP_200_OK,
        )


def read_tabular_file(file_path):
    if not file_path or not os.path.exists(file_path):
        raise FileNotFoundError(f"Ficheiro não encontrado: {file_path}")

    extension = os.path.splitext(file_path)[1].lower()

    if extension == ".csv":
        return read_csv_robust(file_path)

    if extension in {".xlsx", ".xls"}:
        return pd.read_excel(file_path)

    raise ValueError(f"Formato de ficheiro não suportado: {extension}")


def read_csv_robust(file_path):
    attempts = [
        {"sep": None, "engine": "python", "encoding": "utf-8"},
        {"sep": ";", "engine": "python", "encoding": "utf-8"},
        {"sep": ",", "engine": "python", "encoding": "utf-8"},
        {"sep": "\t", "engine": "python", "encoding": "utf-8"},
        {"sep": None, "engine": "python", "encoding": "utf-8-sig"},
        {"sep": ";", "engine": "python", "encoding": "utf-8-sig"},
        {"sep": ",", "engine": "python", "encoding": "utf-8-sig"},
        {"sep": None, "engine": "python", "encoding": "cp1252"},
        {"sep": ";", "engine": "python", "encoding": "cp1252"},
        {"sep": ",", "engine": "python", "encoding": "cp1252"},
    ]

    last_exception = None

    for attempt in attempts:
        try:
            df = pd.read_csv(
                file_path,
                sep=attempt["sep"],
                engine=attempt["engine"],
                encoding=attempt["encoding"]
            )

            if df is not None and len(df.columns) > 1:
                return df

        except (ParserError, UnicodeDecodeError, ValueError) as exc:
            last_exception = exc

    raise ValueError(f"Erro ao ler CSV '{file_path}': {last_exception}")


def normalize_dataframe(df):
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    df = df.replace([np.inf, -np.inf], np.nan)
    df = df.astype(object)
    df = df.where(pd.notnull(df), None)
    return df


def sanitize_records(records):
    sanitized = []

    for record in records:
        clean_record = {}
        for key, value in record.items():
            if pd.isna(value):
                clean_record[key] = None
            elif value == np.inf or value == -np.inf:
                clean_record[key] = None
            else:
                clean_record[key] = value
        sanitized.append(clean_record)

    return sanitized


def build_schedule_data(df):
    df = normalize_dataframe(df)
    return {
        "classes": sanitize_records(df.to_dict(orient="records"))
    }


def build_rooms_data(df):
    df = normalize_dataframe(df)
    return {
        "rooms": sanitize_records(df.to_dict(orient="records"))
    }


class ProblemSolutionsListView(APIView):
    def get(self, request, problem_id):
        problem = get_object_or_404(ProblemDraft, pk=problem_id)
        queryset = problem.solutions.all().order_by("-created_at")
        serializer = SolutionListSerializer(queryset, many=True)
        return Response(serializer.data)


class SolutionDetailView(APIView):
    def get(self, request, problem_id, solution_id):
        solution = get_object_or_404(
            Solution,
            id=solution_id,
            problem_id=problem_id
        )

        serializer = SolutionDetailSerializer(solution, context={"request": request})
        return Response(serializer.data)


class ProblemRoomFeatureResolutionAnalysisView(APIView):
    def get(self, request, problem_id, *args, **kwargs):
        try:
            problem_draft = ProblemDraft.objects.get(pk=problem_id)
        except ProblemDraft.DoesNotExist:
            return Response(
                {"error": "Problem draft not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if not problem_draft.uploaded_schedule:
            return Response(
                {"error": "O problema não tem ficheiro de horário associado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not problem_draft.uploaded_rooms_file:
            return Response(
                {"error": "O problema não tem ficheiro de salas associado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            schedule_df = load_schedule_dataframe(
                problem_draft.uploaded_schedule.file.path
            )
            rooms_df = read_schedule_file(
                problem_draft.uploaded_rooms_file.file.path
            )

            schedule_mapping = (problem_draft.mapping_data or {}).get("mapping", {}) or {}
            rooms_mapping_data = problem_draft.rooms_mapping_data or {}
            existing_resolution = problem_draft.room_feature_resolution or {}

            requested_values = extract_requested_room_feature_values(
                schedule_df,
                schedule_mapping,
            )
            available_room_features = extract_room_features_from_rooms_dataframe(
                rooms_df,
                rooms_mapping_data,
            )

            response_data = build_missing_room_feature_resolution_analysis(
                requested_values=requested_values,
                room_features=available_room_features,
                existing_resolution=existing_resolution,
            )

            return Response(
                {
                    "problem_id": problem_draft.id,
                    "requested_values": response_data["requested_values"],
                    "available_room_features": response_data["available_room_features"],
                    "summary": response_data["summary"],
                },
                status=status.HTTP_200_OK,
            )

        except Exception as exc:
            return Response(
                {
                    "error": "Não foi possível analisar as características em falta.",
                    "details": str(exc),
                    "trace": traceback.format_exc(),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )


def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip().lower()
    if not text:
        return ""

    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("-", "_").replace("/", "_")
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^a-z0-9_]+", "", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def get_column_from_mapping(df, mapping, canonical_key):
    mapped_column = mapping.get(canonical_key)
    if mapped_column and mapped_column in df.columns:
        return mapped_column

    if not mapped_column:
        return None

    normalized_lookup = {
        normalize_text(column): column
        for column in df.columns
    }

    normalized_mapped_column = normalize_text(mapped_column)
    return normalized_lookup.get(normalized_mapped_column)


def split_feature_tokens(value):
    if value is None:
        return []

    text = str(value).strip()
    if not text:
        return []

    parts = re.split(r"[;,|]+", text)
    return [part.strip() for part in parts if part.strip()]


def extract_requested_room_feature_values(schedule_df, schedule_mapping):
    requested_column = get_column_from_mapping(
        schedule_df,
        schedule_mapping,
        "caracteristicas_pedidas_para_sala",
    )

    if not requested_column:
        return []

    seen = set()
    values = []

    for raw_cell_value in schedule_df[requested_column].dropna().tolist():
        tokens = split_feature_tokens(raw_cell_value)

        for token in tokens:
            normalized = normalize_text(token)
            if not normalized or normalized in seen:
                continue

            seen.add(normalized)
            values.append({
                "label": token,
                "normalized": normalized,
            })

    values.sort(key=lambda item: item["normalized"])
    return values


def extract_room_features_from_rooms_dataframe(rooms_df, rooms_mapping_data):
    if rooms_df is None or rooms_df.empty:
        return []

    characteristics = (rooms_mapping_data or {}).get("characteristics", {}) or {}
    char_format = characteristics.get("format")
    config = characteristics.get("config", {}) or {}

    seen = set()
    features = []

    def register_feature(raw_feature):
        normalized = normalize_text(raw_feature)
        if not normalized or normalized in seen:
            return

        seen.add(normalized)
        features.append({
            "label": str(raw_feature).strip(),
            "normalized": normalized,
        })

    if char_format == "single_column_list":
        source_column = config.get("source_column")
        separator = config.get("separator", ",")

        if source_column and source_column in rooms_df.columns:
            for raw_value in rooms_df[source_column].dropna().tolist():
                text = str(raw_value).strip()
                if not text:
                    continue

                parts = [part.strip() for part in text.split(separator) if part.strip()]
                for part in parts:
                    register_feature(part)

    elif char_format == "multiple_columns":
        selected_columns = config.get("selected_columns", []) or []

        for column in selected_columns:
            if column in rooms_df.columns:
                register_feature(column)

    elif char_format == "range_columns":
        start_column = config.get("start_column")
        end_column = config.get("end_column")

        if (
            start_column
            and end_column
            and start_column in rooms_df.columns
            and end_column in rooms_df.columns
        ):
            ranged_df = rooms_df.loc[:, start_column:end_column]
            for column in ranged_df.columns:
                register_feature(column)

    features.sort(key=lambda item: item["normalized"])
    return features


def suggest_room_feature_targets(requested_normalized, available_room_features):
    exact_matches = [
        item for item in available_room_features
        if item["normalized"] == requested_normalized
    ]
    if exact_matches:
        return exact_matches

    contains_matches = [
        item for item in available_room_features
        if requested_normalized in item["normalized"]
        or item["normalized"] in requested_normalized
    ]
    if contains_matches:
        return contains_matches

    requested_parts = set(requested_normalized.split("_"))
    scored = []

    for item in available_room_features:
        target_parts = set(item["normalized"].split("_"))
        overlap = requested_parts.intersection(target_parts)

        if overlap:
            scored.append((len(overlap), item))

    scored.sort(key=lambda entry: (-entry[0], entry[1]["normalized"]))
    return [item for _, item in scored]


def build_missing_room_feature_resolution_analysis(
    requested_values,
    room_features,
    existing_resolution,
):
    available_normalized = {
        item["normalized"]
        for item in room_features
    }

    existing_items = (existing_resolution or {}).get("requested_values", []) or []
    existing_by_normalized = {}

    for item in existing_items:
        source_value = item.get("source_value")
        if not source_value:
            continue
        existing_by_normalized[normalize_text(source_value)] = item

    missing_items = []

    for requested in requested_values:
        normalized = requested["normalized"]

        if normalized in available_normalized:
            continue

        existing_item = existing_by_normalized.get(normalized)

        missing_items.append({
            "source_value": requested["label"],
            "source_value_normalized": normalized,
            "resolution_type": (
                existing_item.get("resolution_type", "unresolved")
                if existing_item
                else "unresolved"
            ),
            "target_values": (
                existing_item.get("target_values", [])
                if existing_item
                else []
            ),
        })

    unresolved_count = len([
        item for item in missing_items
        if item.get("resolution_type") == "unresolved"
    ])

    return {
        "requested_values": missing_items,
        "available_room_features": [item["label"] for item in room_features],
        "summary": {
            "missing_count": len(missing_items),
            "available_count": len(room_features),
            "unresolved_count": unresolved_count,
        },
    }


def apply_room_feature_resolution(df, mapping_data, room_feature_resolution):
    if df is None or df.empty:
        return df

    mapping = (mapping_data or {}).get("mapping", {}) or {}
    requested_column = get_mapped_column(
        df,
        mapping,
        "caracteristicas_pedidas_para_aula"
    )

    if not requested_column or requested_column not in df.columns:
        return df

    resolution_items = (room_feature_resolution or {}).get("requested_values", []) or []
    if not resolution_items:
        return df

    resolution_by_normalized = {}

    for item in resolution_items:
        source_value = item.get("source_value")
        if not source_value:
            continue
        resolution_by_normalized[normalize_text(source_value)] = item

    def resolve_value(raw_value):
        tokens = split_feature_tokens(raw_value)
        if not tokens:
            return []

        resolved_tokens = []

        for token in tokens:
            normalized_token = normalize_text(token)
            resolution = resolution_by_normalized.get(normalized_token)

            if not resolution:
                resolved_tokens.append({
                    "source_value": token,
                    "resolution_type": "original",
                    "target_values": [token],
                })
                continue

            resolution_type = resolution.get("resolution_type", "unresolved")
            target_values = resolution.get("target_values", []) or []

            if resolution_type == "none_required":
                continue

            if resolution_type == "map_to_room_feature":
                resolved_tokens.append({
                    "source_value": token,
                    "resolution_type": "map_to_room_feature",
                    "target_values": target_values,
                })
            elif resolution_type == "no_match":
                resolved_tokens.append({
                    "source_value": token,
                    "resolution_type": "no_match",
                    "target_values": [],
                })
            else:
                resolved_tokens.append({
                    "source_value": token,
                    "resolution_type": "unresolved",
                    "target_values": [],
                })

        return resolved_tokens

    working_df = df.copy()
    working_df["_resolved_requested_room_features"] = working_df[requested_column].apply(resolve_value)

    return working_df
